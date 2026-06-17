"""Create formatted Excel workbooks from extracted DataFrames."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from formula_detector import FormulaRule, detect_formula_rules, formula_for_row, normalize_number


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FORMULA_FILL = PatternFill("solid", fgColor="E2F0D9")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def _safe_sheet_name(name: str, existing: set[str]) -> str:
    """Create a valid, unique Excel worksheet name."""

    cleaned = "".join(char for char in name if char not in r"[]:*?/\\")[:31] or "Table"
    candidate = cleaned
    counter = 2
    while candidate in existing:
        suffix = f"_{counter}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        counter += 1
    existing.add(candidate)
    return candidate


def _coerce_cell_value(value: object) -> object:
    """Keep text headers as text while writing numeric-looking values as numbers."""

    if pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return None
    cleaned = (
        text.replace(",", "")
        .replace("\u20b9", "")
        .replace("Rs.", "")
        .replace("INR", "")
        .replace("%", "")
        .strip()
    )
    try:
        return float(cleaned)
    except ValueError:
        # Extracted engineering notes often start with "=" but are explanatory
        # text, not workbook formulas. Prefix them so Excel displays the note.
        if text.startswith("="):
            return f"'{text}"
        return text


def _row_supports_formula(row: pd.Series, rule: FormulaRule) -> bool:
    """Only apply a formula where the target and source cells are numeric values."""

    relevant_columns = (rule.target_col, *rule.source_cols)
    for column_index in relevant_columns:
        if normalize_number(row.iloc[column_index]) is None:
            return False
    return True


def _apply_table_formatting(ws, row_count: int, col_count: int, formula_cols: set[int]) -> None:
    """Apply basic readable formatting to a worksheet."""

    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row in ws.iter_rows(min_row=2, max_row=row_count + 1, max_col=col_count):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if cell.column - 1 in formula_cols:
                cell.fill = FORMULA_FILL

    for column_index in range(1, col_count + 1):
        letter = get_column_letter(column_index)
        max_length = 10
        for cell in ws[letter]:
            max_length = max(max_length, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max_length + 2, 42)

    if row_count > 0 and col_count > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}{row_count + 1}"


def write_dataframe_to_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    sheet_name: str,
    existing_names: set[str],
) -> list[FormulaRule]:
    """Write one DataFrame to one worksheet and insert detected formulas."""

    ws = wb.create_sheet(_safe_sheet_name(sheet_name, existing_names))
    rules = detect_formula_rules(df)
    rule_by_target = {rule.target_col: rule for rule in rules}

    for column_index, column_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=column_index, value=str(column_name))

    for row_offset, (_, row) in enumerate(df.iterrows(), start=2):
        for column_index, value in enumerate(row, start=1):
            rule = rule_by_target.get(column_index - 1)
            if rule and _row_supports_formula(row, rule):
                ws.cell(row=row_offset, column=column_index, value=formula_for_row(rule, row_offset))
            else:
                ws.cell(row=row_offset, column=column_index, value=_coerce_cell_value(value))

    _apply_table_formatting(ws, len(df), len(df.columns), set(rule_by_target))
    return rules


def generate_workbook(tables: list[tuple[pd.DataFrame, str]]) -> tuple[bytes, dict[str, list[FormulaRule]]]:
    """Generate an XLSX workbook and return bytes plus detected formula metadata."""

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    existing_names: set[str] = set()
    formula_summary: dict[str, list[FormulaRule]] = {}

    if not tables:
        ws = wb.create_sheet("No Tables")
        ws["A1"] = "No tables were extracted from the PDF."
        ws["A1"].font = Font(bold=True)
    else:
        for index, (df, source_name) in enumerate(tables, start=1):
            sheet_name = f"Table {index}"
            rules = write_dataframe_to_sheet(wb, df, sheet_name, existing_names)
            formula_summary[f"{sheet_name} ({source_name})"] = rules

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), formula_summary
